from openpyxl import load_workbook


class Saver:
    def __init__(self, filename):
        self.filename = filename
        self.work_book = load_workbook(self.filename)
        self.sheet = self.work_book.active  # do active one page

    def add_on_top(self, data):
        # append don't work here, I don't know why
        self.sheet.insert_rows(2)  # add row on the top
        for cols, value in enumerate(data):
            self.sheet.cell(row=2, column=cols + 1).value = value

        self.work_book.save(self.filename)
